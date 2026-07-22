# -*- coding: utf-8 -*-
"""
现券预录单要素识别 · 正则提取脚本
思路：切段(我方/对方) + 白名单认领我方 + 字段正则 + 拆单展开
用法：python3 提取脚本.py  → 生成 现券预录单要素_脚本输出.xlsx
"""
import re, glob, os
from datetime import date

import counterparty_fields as counterparty
from counterparty_fields import (
    DEALER_NAME,
    RE_ORG,
    clean_counterparty_candidate as _clean_counterparty_candidate,
    dealer_name_from_code,
    extract_full_account_from_line,
    extract_labeled_account,
    extract_role_account,
    extract_subject_name,
    find_known_dealer_code,
    is_valid_counterparty_account,
    looks_like_account_name,
    looks_like_plain_org,
    normalize_visible_org,
    pick_org,
    split_org_from_account,
)

OUTPUT_COLUMNS = [
    "市场","交易方向","交易日期","债券代码","债券简称","到期收益率","行权收益率",
    "原始净价","交易净价","交易规模万","我方账户","对方账户","过券","中介","中介费",
    "对手方交易员","清算速度","约定号","对手方交易员代码","对手方交易商代码",
    "对手方交易商简称","对手方交易主体代码","报价发起方","备注","原文"
]

# ========== 1. 我方账户白名单（按实际管理产品维护）==========
WHITELIST = [
    "中信信托信昱11号","中信信托信昱13号","中信信托华盈添利4号",
    "粤财信托添添益1号","粤财信托锐益2号","粤财信托广盈恒益1号",
    "财信信托湘信财盈73号","财信信托湘信财盈74号",
    "中粮佳盈1号","广粤尊享77号",
]
MINE_PREFIXES = ("中信信托","粤财信托","财信信托")


# 全角 ASCII（数字/字母/标点，U+FF01~U+FF5E）→半角：文本经不同系统/输入法转发后
# 常出现全角数字债券代码、全角加号/冒号等，统一转半角后下游正则不用逐个再兼容一遍全角变体。
_FULLWIDTH_TRANS = str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(0x7E - 0x21 + 1)})


def _norm_basic(s):
    s = (s or "").translate(_FULLWIDTH_TRANS)
    s = s.replace("【", " ").replace("】", " ").replace("（", "(").replace("）", ")")
    return re.sub(r'[\u3000\xa0]+', ' ', s)


def _compact_text(s):
    return re.sub(r'\s+','', _norm_basic(s))


def _mine_aliases(name):
    full = _compact_text(name)
    aliases = {full}
    core = full
    for prefix in MINE_PREFIXES:
        if core.startswith(prefix):
            core = core[len(prefix):]
            aliases.add(core)
            break
    aliases.add(core.rstrip("号"))
    if core and not core.endswith("号"):
        aliases.add(core + "号")
    return {x for x in aliases if len(x) >= 3}


MINE_ALIAS_TO_CANON = {}
for _name in WHITELIST:
    for _alias in _mine_aliases(_name):
        MINE_ALIAS_TO_CANON.setdefault(_alias, _name)
MINE_ALIASES = sorted(MINE_ALIAS_TO_CANON, key=len, reverse=True)

# ========== 2. 字段正则 ==========
RE_BOND   = re.compile(
    r'(?:(\d{6})\s*[.\-/]?\s*(SH|SZ)|(SH|SZ)\s*[:.\-/]?\s*(\d{6}))',
    re.I,
)  # 兼容 245555.SH / 245555-SH / 245555 SH / SH:245555
RE_PRICE  = re.compile(r'净价\s*(\d{2,3}(?:\.\d+)?)')
RE_PRICE2 = re.compile(r'(?<!\d)(1\d{2}(?:\.\d+)?)(?!\d)')      # 兜底 100~199
RE_PRICE3 = re.compile(r'\d+(?:\.\d+)?%\s*(\d{2,3}(?:\.\d+)?)')  # 收益率后紧跟净价，如 2.564% 99.814
RE_ORIG_PRICE = re.compile(r'(?:原始净价|委托净价)\s*[:：]?\s*(\d{2,3}(?:\.\d+)?)')
RE_DEAL_PRICE = re.compile(r'(?:交易净价|成交净价)\s*[:：]?\s*(\d{2,3}(?:\.\d+)?)')
RE_YIELD  = re.compile(r'(\d+(?:\.\d+)?)\s*%')
RE_YIELD2 = re.compile(r'(?<![\d.])([12]\.\d{2,3})(?![\d%])')   # 裸 1.xx/2.xx
RE_MATURITY_YIELD = re.compile(r'(?<!行权)(?:(?:到期)?收益率|YTM)\s*[:：]?\s*(\d+(?:\.\d+)?)\s*(%)?', re.I)
RE_EXERCISE_YIELD = re.compile(r'(?:行权收益率|行权YTM|行权)\s*[:：]?\s*(\d{1,2}(?:\.\d+)?)(?![\d.])\s*(%)?', re.I)
RE_BOND_NAME_LABEL = re.compile(r'(?:债券简称|券简称|债券名称)\s*[:：]?\s*([A-Za-z0-9一-龥－\-]+)')
RE_DATEF  = re.compile(r'(20\d{2})\s*[-./]\s*(\d{1,2})\s*[-./]\s*(\d{1,2})')
RE_DATEZH = re.compile(r'(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?')
RE_DATES  = re.compile(r'(?<!\d)(\d{1,2})[.\-/](\d{1,2})\s*交易所')
RE_DATEC  = re.compile(r'(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)')
RE_DATES2 = re.compile(r'(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)')
RE_DATE4T = re.compile(r'(?<!\d)(\d{2})(\d{2})\s*交易')
RE_DATE4P = re.compile(r'(?<!\d)(\d{2})(\d{2})(?=\s*[+＋]\d{1,2}(?!\d))')
RE_DATEM  = re.compile(r'(?<!\d)(\d{1,2})月(\d{1,2})日')
# 清算速度："今日+0"/"上交所+0"/"6/30+0" 这类写法里，紧跟在日期/交易所关键词后面、
# 中间不隔其他数字的 "+N" 才是清算速度；像 "5+5"(期限写法) 前面不是日期/关键词，不会误中。
RE_SETTLE_TODAY = re.compile(r'今[日天][ ]*[+＋](\d{1,2})(?!\d)')
RE_SETTLE_EXCH  = re.compile(r'(?:上交所|深交所|交易所)[ ]*[+＋](\d{1,2})(?!\d)')
RE_SETTLE_MD    = re.compile(r'(?<!\d)(?:20\d{2}[./-])?\d{1,2}[./-]\d{1,2}[ ]*[+＋](\d{1,2})(?!\d)')
RE_SETTLE_8D    = re.compile(r'(?<!\d)20\d{6}[ ]*[+＋](\d{1,2})(?!\d)')
RE_SETTLE_4D    = re.compile(r'(?<!\d)\d{4}[ ]*[+＋](\d{1,2})(?!\d)')
RE_SETTLE_T     = re.compile(r'(?<![A-Za-z0-9])T\s*[+＋]\s*(\d{1,2})(?!\d)', re.I)
RE_T8_MARKER    = re.compile(
    r'^(?:【?[^\n]{0,24}(?:深交所(?:交易)?信息|深交所要素)】?|[^\n]{0,24}深交所(?:交易)?信息[:：]?)\s*$',
    re.M,
)
RE_YD     = re.compile(r'约(?:定号)?[：:\s]*([0-9]+(?:\s*[+＋、，,]\s*[0-9]+)*)')  # 支持 + 、 ，分隔及"约"简写
RE_YD_ALL = re.compile(r'\b(\d{3,10})\b')
RE_YD_LINE= re.compile(r'约(?:定号)?[：:\s]*([0-9]{3,10}(?:\s*[+＋、，,]\s*[0-9]{3,10})*)')
RE_DEALER = re.compile(r'交易商(?:代码|号)?\s*[:：]?\s*(\d{6})')
RE_TRADER = re.compile(r'交易员(?:代码|号)?\s*[:：]?\s*([0-9A-Z]{8})(?![0-9A-Z])', re.I)
RE_TRADER2= re.compile(r'交易员(?:及交易员代码|代码|名称|号)?\s*[：:]?\s*[一-龥]{2,4}\s*[（(]?\s*([0-9A-Z]{7,8})(?![0-9A-Z])', re.I)  # 名在码前：赵越 00H00010 / 付玉 007Z0039
RE_SUBJ   = re.compile(r'(?:交易(?:商)?主体(?:代码)?)\s*[:：]?\s*(36\d{8})')
RE_SUBJNM = re.compile(r'交易主体名称\s*[:：]?\s*([一-龥A-Z0-9\-（）()]+)')
# 我方经纪主体身份的固定四件套(交易商代码+交易员代码+交易主体代码+以"机构经纪"结尾的主体名称)，
# 顺序固定、标签固定，可以跨行/跨空白匹配；命中即代表这一整段是"我方"的席位信息，不是对方账户。
# 四件套之间允许夹杂"交易商名称/交易员名称/交易主体类型/交易主体简称"这类补充标签(有的记录会展开写)，
# 但间隔要卡得比较紧(<=30字符)：真正的对方信息(主体码+产品名+规模+约定号)通常远超这个长度，
# 卡紧了才不会把"交易商代码：xxx"(对方的)当成起点，一路跨过对方整段信息连到我方那句上
RE_SELF_BROKER_BLOCK = re.compile(
    r'交易商(?:代码|号)?\s*[：:]?\s*\d{6}[\s\S]{0,30}?'
    r'交易员(?:代码|号)?\s*[：:]?\s*[0-9A-Z]{7,8}[\s\S]{0,30}?'
    r'交易(?:商)?主体代码\s*[：:]?\s*(36\d{8})[\s\S]{0,30}?'
    r'交易(?:商)?主体(?:全称|名称)\s*[：:]\s*[一-龥]{2,16}机构经纪',
    re.I,
)
RE_ICODE  = re.compile(r'i\d{9}', re.I)
RE_PERSON = re.compile(r'[（(]([一-龥]{2,4})[)）]')
RE_INIT   = re.compile(r'(卖方发单|买方发单|卖家先发|买家先发|卖出先发|买入先发)')
RE_MID    = re.compile(r'中介费\s*[:：]\s*(\d+(?:\.\d+)?)')
# 规模数字：支持千分位逗号写法，如 "1,000w"（先试逗号分组，再退化到普通整数/小数）
_SIZE_NUM = r'\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+(?:\.\d+)?'
RE_SIZEU  = re.compile(r'(?<![A-Za-z0-9.])(' + _SIZE_NUM + r')\s*([wW万]元?)(?![A-Za-z0-9])')
RE_SIZEK  = re.compile(r'(?<![A-Za-z0-9.])(' + _SIZE_NUM + r')\s*([kK])(?:\s*[wW])?元?(?![A-Za-z0-9])')
RE_SIZEE  = re.compile(r'(?<![A-Za-z0-9.])(' + _SIZE_NUM + r')\s*([eE亿]元?)(?![A-Za-z0-9])')
RE_SIZE_TOKEN = re.compile(r'(?<![A-Za-z0-9.])(' + _SIZE_NUM + r')\s*(亿元?|[eE]元?|[kK](?:\s*[wW])?元?|[wW万]元?)(?![A-Za-z0-9])')
MKT = {"SH":"上交所","SZ":"深交所"}
# 我方产品模式（白名单兜底）：允许产品名里夹空格，如"粤财信托添添益 1号"
RE_MINE_PAT = re.compile(r'((?:中信信托|粤财信托|财信信托)[一-龥A-Za-z0-9 ]{1,18}?号)')

RE_SIZE_NJ = re.compile(r'(?<![.\d])(\d{3,6})\s*净价')          # 3000 净价100（前不接小数点，避免2.025误取025）
RE_SIZE_NJ2= re.compile(r'(\d{3,6})\s+\d+(?:\.\d+)?\s*净价')  # 5000 1.98 净价
RE_SIZE_PR = re.compile(r'(\d{3,6})\s+1\d{2}(?:\.\d+)?(?!\d)')  # 1000 100.001
RE_DEALER_ORG = re.compile(r'交易商(?:代码|号)?\s*[:：]?\s*\d{6}\s*[（(]?\s*([一-龥]{2,10}(?:证券资管|证券自营|证券机构经纪|证券|资管|基金|信托|银行|期货))')
RE_DIRECTION_LABEL = re.compile(r'(?:交易)?方向\s*[:：]?\s*(买入|卖出)')
RE_TRADE_CONTEXT = re.compile(r'净价|到期|行权|出给|卖给|买自|\bto\b|\bfrom\b|买入|卖出|卖方|买方', re.I)


def _bond_parts(match):
    """把代码前缀/后缀的多种写法统一成 (6位代码, SH/SZ)。"""
    if not match:
        return "", ""
    digits = match.group(1) or match.group(4)
    exchange = (match.group(2) or match.group(3) or "").upper()
    return digits, exchange


def _canonical_bond(match):
    digits, exchange = _bond_parts(match)
    return f"{digits}.{exchange}" if digits and exchange else ""

# ========== 3. 主执行链路 ==========
def parse_text(text):
    text=text or ""
    self_subjects=self_broker_subjects(text)
    allrows=[]
    for rec in split_records(text):
        allrows.extend(parse_record(rec, self_subjects=self_subjects))
    # 未命中白名单的行（多为对手只报券商席位、我方仅以席位i码出现）→ 打备注提示人工确认
    for r in allrows:
        if str(r.get("我方账户","")).startswith("(未") and not r.get("备注"):
            r["备注"]="未命中我方白名单：我方仅以席位i码出现，请人工确认我方产品及债券简称"
    return allrows

def split_records(text):
    raw=[]
    for chunk in re.split(r'\n\s*\n', text):
        if looks_like_trade_chunk(chunk):
            raw.append(chunk)
        elif raw and chunk.strip():        # 无债券代码=上一条的续接(约定号/卖出方等)，用换行连接保留结构
            raw[-1]+='\n'+chunk
    # 二次切
    recs=[]
    carry_prefix=""
    carry_bond_code=""
    for chunk in raw:
        lines=[l for l in chunk.split('\n') if l.strip()]
        first_bond_line=next((i for i,l in enumerate(lines) if RE_BOND.search(l)), -1)
        first_bond_code=""
        if first_bond_line >= 0:
            m=RE_BOND.search(lines[first_bond_line])
            if m: first_bond_code=_canonical_bond(m)
        if first_bond_line > 0:
            carry_prefix='\n'.join(lines[:first_bond_line])
            carry_bond_code=first_bond_code
        elif carry_prefix and first_bond_code and first_bond_code == carry_bond_code:
            chunk=carry_prefix + '\n' + chunk
            lines=[l for l in chunk.split('\n') if l.strip()]
        elif first_bond_code and carry_bond_code and first_bond_code != carry_bond_code:
            carry_prefix=""
            carry_bond_code=""
        # a) 每行都自成一笔(各含债券代码+约定号)→按行拆(如上交所两笔挨着无空行)
        if len(lines)>1 and all(RE_BOND.search(l) and RE_YD.search(l) for l in lines):
            recs.extend(lines); continue
        # b) 块内多个"各自带出给/买入的完整交易"(方向标记在每段内)→按债券代码切；
        #    "共享抬头+多只券"(段内无方向标记，如 华创 出给 财信 后跟两只券)不切，留给 multibond
        codes=[mm.start() for mm in RE_BOND.finditer(chunk)]
        if len(codes)>=2:
            # 按"债券代码所在行的行首"切分，而不是代码本身的字符位置：
            # 简称写在代码前面时（如"26中财G3 245168.SH"），代码前的简称/日期跟代码同属一行，
            # 按代码字符位置切分会把这一行的简称丢给上一段、又把下一行的简称错吞进本段，
            # 导致债券简称跟债券代码错位。按行首切分能保住"简称+代码"同行的完整性。
            line_starts=[chunk.rfind('\n', 0, c) + 1 for c in codes]
            header=chunk[:line_starts[0]]   # 首个代码所在行之前的共享抬头(我方账户/方向等)，每段都要回填
            segs=[chunk[line_starts[i]:line_starts[i+1]] for i in range(len(line_starts)-1)]+[chunk[line_starts[-1]:]]
            if all(re.search(r'出给|\bto\b|from|买入|卖出|卖方|买方', s) for s in segs):
                if header.strip():
                    segs=[header+'\n'+s for s in segs]
                recs.extend(segs); continue
        recs.append(chunk)
    return recs

def looks_like_trade_chunk(chunk):
    if RE_BOND.search(chunk):
        return True
    # 无债券代码时，"约定号/主体代码"这类续行信息经常单独成段，不能仅凭一个弱候选简称就把它误判成新交易。
    # 这里要求同时满足"核心交易语境" + "高置信债券简称"，把补充说明块留给上一笔记录拼接。
    if not RE_TRADE_CONTEXT.search(chunk):
        return False
    return has_confident_bond_name(chunk)

def self_broker_subjects(text):
    """扫描整份原始文本，收集"重复出现"的我方经纪席位占位码。
    "交易主体名称：...机构经纪"这个后缀本身不是我方专属的——语料里也见过对手方自己的
    经纪主体同样带"机构经纪"后缀（比如对手方是国信证券，也会给一个机构经纪代码，但那个
    代码全篇只出现一次，跟这一笔对手方绑死）。真正能区分"这是我方固定占位码"的信号是
    重复：我方每次交易不管对手是谁，报的都是同一个交易商+交易员+主体代码三元组，
    所以它会在同一份文本里反复出现；对手方自己的机构经纪代码只会跟那一笔对手绑在一起，
    只出现一次。因此只把出现 >=2 次的主体代码当成我方占位码，只出现一次的一律当对方的。"""
    from collections import Counter
    counts = Counter(m.group(1) for m in RE_SELF_BROKER_BLOCK.finditer(text))
    return {code for code, n in counts.items() if n >= 2}

def parse_record(rec, self_subjects=frozenset()):
    ctx = build_parse_context(rec, self_subjects=self_subjects)
    if not ctx:
        return []
    template_id = _detect_template_from_ctx(ctx)
    ctx["template_id"] = template_id
    handler_map = {
        "T1": parse_template_single,
        "T2": parse_template_single,
        "T3": parse_template_t3,
        "T4": parse_template_t4,
        "T5": parse_template_t5,
        "T6": parse_template_t6,
        "T7": parse_template_t7,
        "T8": parse_template_t8,
    }
    return handler_map.get(template_id, parse_template_single)(ctx)

# ========== 4. 上下文构建 ==========
def build_parse_context(rec, self_subjects=frozenset()):
    rec = re.sub(r'[ \t]+', ' ', _norm_basic(rec).strip())
    main_rec, supplement = split_main_and_supplement(rec)
    base_rec = main_rec or rec
    bm = RE_BOND.search(base_rec) or RE_BOND.search(rec)
    raw_code = bm.group(0) if bm else ""
    code = _canonical_bond(bm)
    _, exchange = _bond_parts(bm)
    mkt = MKT.get(exchange, "") if bm else infer_market(rec)
    labeled_name = RE_BOND_NAME_LABEL.search(base_rec) or RE_BOND_NAME_LABEL.search(rec)
    name = labeled_name.group(1) if labeled_name else (get_name(base_rec, raw_code) or get_name(rec, raw_code))
    if not (code or name):
        return None

    date = norm_date(rec)
    settle_speed = ""
    sm = RE_SETTLE_T.search(rec) or RE_SETTLE_TODAY.search(rec)
    if sm:
        settle_speed = f"T+{sm.group(1)}"
        if not date and RE_SETTLE_TODAY.search(rec):
            date = _today_iso()
    else:
        sm = RE_SETTLE_EXCH.search(rec) or RE_SETTLE_MD.search(rec) or RE_SETTLE_8D.search(rec) or RE_SETTLE_4D.search(rec)
        if sm:
            settle_speed = f"T+{sm.group(1)}"
            if not date and RE_SETTLE_EXCH.search(rec):
                date = _today_iso()

    price = ""
    pm = RE_PRICE.search(base_rec)
    if pm:
        price = pm.group(1)
    else:
        pm2 = RE_PRICE2.search(base_rec.replace(code, ''))
        if pm2:
            price = pm2.group(1)
        else:
            pm3 = RE_PRICE3.search(base_rec)
            if pm3:
                price = pm3.group(1)
    xk = RE_EXERCISE_YIELD.search(base_rec)
    if xk:
        xingquan = xk.group(1) + ('%' if xk.group(2) else '')
    else:
        xk = re.search(r'行权[^\d]{0,3}(\d+\.\d+)|(\d+\.\d+)\s*行权', base_rec)
        xingquan = (xk.group(1) or xk.group(2)) if xk else ""
    ym = RE_MATURITY_YIELD.search(base_rec)
    if ym:
        yld = ym.group(1) + ('%' if ym.group(2) else '')
    else:
        yld = ""
        for pct in RE_YIELD.finditer(base_rec):
            if "行权" not in base_rec[max(0, pct.start() - 8):pct.start()]:
                yld = pct.group(1) + '%'
                break
    dqk = re.search(r'(\d+(?:\.\d+)?)\s*到期', base_rec)
    if dqk and not yld:
        yld = dqk.group(1)
    init_raw = RE_INIT.search(rec)
    init_raw = init_raw.group(1) if init_raw else ""
    mid = RE_MID.search(rec)
    midfee = mid.group(1) if mid else ""

    mine, mpos = find_mine(base_rec)
    direction, mineseg, otherseg = direction_and_split(base_rec, mpos)
    counter_seg = (otherseg.rstrip() + "\n" + supplement.lstrip()).strip() if supplement else otherseg
    if mine and mpos >= 0:
        win = base_rec[mpos:mpos + 90]
        if re.search(r'卖出|卖方|卖家', win):
            direction = '卖出'
        elif re.search(r'买入方|买家', win):
            direction = '买入'
    if not mine:
        mine, mpos = find_mine(rec)
    if not mine:
        mine = "(未命中白名单)"

    initiator = ""
    if init_raw:
        seller_first = init_raw in ("卖方发单", "卖家先发", "卖出先发")
        if direction == "卖出":
            initiator = "我方发起" if seller_first else "对方发起"
        elif direction == "买入":
            initiator = "对方发起" if seller_first else "我方发起"
    if not initiator:
        m_direct_init = re.search(r'(我方|对方|我司|对手方?)\s*发(?![一-龥])', rec)
        if m_direct_init:
            initiator = "我方发起" if m_direct_init.group(1) in ("我方", "我司") else "对方发起"
    if not initiator:
        m_bare_init = re.search(RE_ORG.pattern + r'\s*发(?![一-龥])', rec)
        if m_bare_init:
            hit = m_bare_init.group(0)
            if counter_seg and hit in counter_seg:
                initiator = "对方发起"
            elif mineseg and hit in mineseg:
                initiator = "我方发起"

    counter = resolve_counterparty(counter_seg, rec=rec, mine=mine, self_subjects=self_subjects)
    counter_dealer_code = counter["dealer_code"]
    counter_trader_code = counter["trader_code"] if mkt == "深交所" else counter["icode"]
    counter_subj = counter["subj_code"]
    counter_short = counter["short"]
    counter_acct = counter["account"]
    o_person = counter["person"]
    if mkt == "上交所":
        counter_dealer_code = ""
        counter_short = ""
        counter_subj = ""
    guoquan = counter["guoquan"]
    if not guoquan and counter_acct and looks_like_plain_org(counter_acct):
        guoquan = counter_acct

    orig_match = RE_ORIG_PRICE.search(base_rec)
    deal_match = RE_DEAL_PRICE.search(base_rec)
    orig = orig_match.group(1) if orig_match else price
    deal = deal_match.group(1) if deal_match else price
    if midfee and price and not deal_match:
        try:
            p = float(price)
            f = float(midfee)
            deal = str(p - f) if direction == "卖出" else str(p + f)
        except:
            pass

    return dict(
        rec=rec,
        base_rec=base_rec,
        supplement=supplement,
        raw_code=raw_code,
        code=code,
        mkt=mkt,
        name=name,
        date=date,
        settle_speed=settle_speed,
        price=price,
        yld=yld,
        xingquan=xingquan,
        init_raw=init_raw,
        midfee=midfee,
        mine=mine,
        mpos=mpos,
        direction=direction,
        mineseg=mineseg,
        otherseg=otherseg,
        counter_seg=counter_seg,
        counter=counter,
        counter_dealer_code=counter_dealer_code,
        counter_trader_code=counter_trader_code,
        counter_subj=counter_subj,
        counter_short=counter_short,
        counter_acct=counter_acct,
        o_person=o_person,
        guoquan=guoquan,
        orig=orig,
        deal=deal,
        initiator=initiator,
        self_subjects=self_subjects,
        table=extract_table(base_rec),
        multibond=extract_multibond(base_rec),
        paren=extract_paren_products(base_rec),
        yds=extract_all_yds(base_rec),
        sizes=expand_splits(base_rec)[1],
    )

def resolve_counterparty(otherseg, rec="", mine="", self_subjects=frozenset()):
    dealer_match = RE_DEALER.search(otherseg)
    o_dealer = dealer_match.group(1) if dealer_match else find_known_dealer_code(otherseg)
    o_trader = (RE_TRADER.search(otherseg) or RE_TRADER2.search(otherseg))
    o_trader = o_trader.group(1).upper() if o_trader else ""
    # 主体代码：优先取"裸代码"(不要求紧跟标签)，第一个出现的、且不是已知我方经纪席位码的
    # 36 开头代码就是对方的——有些记录里"交易主体代码："标签本身是空的，真正的代码单独另起
    # 一行(无标签)，反而是本方经纪身份块里的代码带着完整标签，只认标签会取反
    o_subj=""
    for m in re.finditer(r'(36\d{8})', otherseg):
        if m.group(1) not in self_subjects:
            o_subj=m.group(1); break
    if not o_subj:
        m=RE_SUBJ.search(otherseg)
        if m and m.group(1) not in self_subjects:
            o_subj=m.group(1)
    o_icode  = RE_ICODE.search(otherseg).group(0).lower() if RE_ICODE.search(otherseg) else ""
    o_person = person_name(otherseg)
    o_subjname = extract_subject_name(otherseg)
    o_prod   = pick_prod(otherseg)
    o_labeled = extract_role_account(otherseg, "counter") or extract_role_account(rec, "counter")
    o_head   = extract_counterparty_head(otherseg)
    role_account_org, role_execution_org = counterparty.extract_account_execution_orgs(
        otherseg, mine_related_text
    )
    if role_account_org:
        o_head = role_account_org
    acct = resolve_counterparty_account(
        labeled=o_labeled,
        subject_name=o_subjname,
        product=o_prod,
        head=o_head,
    )
    org = resolve_guoquan(
        otherseg,
        dealer_code=o_dealer,
        subject_name=o_subjname,
        head=o_head,
        account=acct,
        execution_org=role_execution_org,
    )
    if not acct and org:
        # 对方只报机构、未报产品账户时，机构本身可作为对方账户；动作词和标签噪声不会进入这里。
        acct = org
    return dict(
        dealer_code=o_dealer,
        trader_code=o_trader,
        subj_code=o_subj,
        icode=o_icode,
        person=o_person,
        subject_name=o_subjname,
        account=acct,
        guoquan=org,
        short=org or dealer_name_from_code(o_dealer),
    )

def expand_splits(rec):
    """返回 [(规模万, 约定号)] 列表；支持 + 连写"""
    yds=extract_all_yds(rec)
    # 规模：找 a+b+c（含括号内 (1000+1000)）——先去掉"约定号 XX+YY"整串，避免把约定号当规模
    rec_wo_yd=RE_YD.sub(' ', rec)
    sizes=[]
    plus=re.search(r'(\d{2,6}(?:\s*[+＋]\s*\d{2,6})+)', rec_wo_yd)
    if plus:
        sizes=[int(x) for x in re.split(r'[+＋]', plus.group(1))]
    return yds, sizes

# ========== 5. 模板识别 ==========
def detect_template(rec, self_subjects=frozenset()):
    ctx = build_parse_context(rec, self_subjects=self_subjects)
    if not ctx:
        return ""
    return _detect_template_from_ctx(ctx)

def _detect_template_from_ctx(ctx):
    if ctx["supplement"]:
        return "T8"
    if ctx["table"]:
        return "T6"
    if ctx["multibond"]:
        return "T7"
    if _looks_like_t4(ctx):
        return "T4"
    if _looks_like_t5(ctx):
        return "T5"
    if _looks_like_t3(ctx):
        return "T3"
    if _looks_like_t2(ctx):
        return "T2"
    return "T1"

def _count_subject_rows(seg):
    return sum(1 for line in (seg or "").splitlines() if re.search(r'36\d{8}', line))

def _count_row_yd_lines(seg):
    return sum(1 for line in (seg or "").splitlines() if RE_YD_LINE.search(line))

def _count_detailed_subject_rows(seg):
    out = 0
    for raw in (seg or "").splitlines():
        line = _strip_leading_enum(raw)
        if not re.search(r'36\d{8}', line):
            continue
        if RE_YD_LINE.search(line) or RE_SIZE_TOKEN.search(line):
            out += 1
            continue
        body = re.sub(r'36\d{8}', ' ', line)
        body = RE_YD.sub(' ', body)
        body = re.sub(r'\d+\.\d+', ' ', body)
        if re.search(r'(?<![\d.])\d{2,6}(?![\d.])', body):
            out += 1
    return out

def _count_account_size_rows(seg):
    return len(_extract_sh_account_rows(seg))

def _has_split_yd(rec):
    return bool(re.search(r'约(?:定号)?[：:\s]*\d+(?:\s*[+＋、，,]\s*\d+)+', rec or ""))

def _has_split_size(rec):
    rec_wo_yd = RE_YD.sub(' ', rec or "")
    return bool(
        re.search(r'\d{2,6}\s*[+＋]\s*\d{2,6}', rec_wo_yd)
        or re.search(r'[（(]\s*\d{2,6}\s*[+＋]\s*\d{2,6}', rec_wo_yd)
    )


RE_SIZE_ONLY_LINE = re.compile(r'(?m)^[ \t]*(' + _SIZE_NUM + r')\s*(亿元?|[eE]元?|[kK](?:\s*[wW])?元?|[wW万]元?)\s*$')


def _structured_counterparty_anchor_positions(seg):
    text = _norm_basic(seg or "")
    if not text:
        return []
    anchor_groups = [
        [m.start() for m in RE_SIZE_ONLY_LINE.finditer(text)],
        [m.start() for m in re.finditer(r'(?m)^[ \t]*交易主体(?:全称|名称|简称)?\s*[：:]\s*[^\n]+$', text)],
        [m.start() for m in re.finditer(r'(?m)^[ \t]*(?:交易主体代码\s*[：:]\s*)?36\d{8}(?:\s|$)', text)],
    ]
    for positions in anchor_groups:
        uniq = sorted(set(positions))
        if len(uniq) >= 2:
            return uniq
    return []


def _has_structured_counterparty_blocks(seg):
    return len(_structured_counterparty_anchor_positions(seg)) >= 2

def _looks_like_t4(ctx):
    if ctx["mkt"] != "上交所":
        return False
    if _count_subject_rows(ctx["counter_seg"]):
        return False
    if _count_row_yd_lines(ctx["counter_seg"]) >= 2:
        return True
    if _count_account_size_rows(ctx["counter_seg"]) >= 2:
        return True
    return False

def _looks_like_t5(ctx):
    seg = ctx["counter_seg"]
    if _has_structured_counterparty_blocks(seg):
        return True
    if re.search(r'交易主体代码/简称/量/约定号', seg):
        return True
    if _count_detailed_subject_rows(seg) >= 2:
        return True
    if ctx["mkt"] == "深交所" and _count_subject_rows(seg) >= 2 and re.search(r'交易主体[:：]\s*[^\n]+', seg):
        return True
    return False

def _looks_like_t3(ctx):
    if ctx["paren"]:
        return True
    if _has_split_size(ctx["base_rec"]):
        return True
    if _has_split_yd(ctx["base_rec"]) or len(ctx["yds"]) > 1:
        return True
    return False

def _looks_like_t2(ctx):
    text = ctx["rec"]
    if re.search(r'交易商(?:代码|号)|交易员(?:代码|号)|交易主体(?:代码|名称|简称|全称)', text):
        return True
    return False

def _looks_like_multi_t8(ctx):
    return bool(
        ctx["paren"]
        or _has_structured_counterparty_blocks(ctx["counter_seg"])
        or len(ctx["yds"]) > 1
        or _count_row_yd_lines(ctx["counter_seg"]) >= 2
        or _count_subject_rows(ctx["counter_seg"]) >= 2
    )

# ========== 6. 模板解析 ==========
def _make_row(ctx, size, yd, acct=None, subj=None):
    return dict(
        市场=ctx["mkt"],
        交易方向=ctx["direction"],
        交易日期=ctx["date"],
        债券代码=ctx["code"],
        债券简称=ctx["name"],
        到期收益率=ctx["yld"],
        行权收益率=ctx["xingquan"],
        原始净价=ctx["orig"],
        交易净价=ctx["deal"],
        交易规模万=size,
        我方账户=ctx["mine"],
        对方账户=acct if acct else ctx["counter_acct"],
        过券=ctx["guoquan"],
        中介="无" if not ctx["midfee"] else "",
        中介费=ctx["midfee"] if ctx["midfee"] else "无",
        对手方交易员=ctx["o_person"],
        清算速度=ctx["settle_speed"],
        约定号=yd,
        对手方交易员代码=ctx["counter_trader_code"],
        对手方交易商代码=ctx["counter_dealer_code"],
        对手方交易商简称=ctx["counter_short"],
        对手方交易主体代码=subj if subj else ctx["counter_subj"],
        报价发起方=ctx["initiator"],
        原文=ctx["rec"],
        备注="",
    )

def _resolve_multi_rows(ctx, allow_subject_anchor=False):
    # 多笔优先级固定为：结构块解析 -> 旧专用规则 -> 约定号锚点兜底 -> 无约定号主体锚点。
    # 这样模板判定依赖结构，不让样本内正则先抢走主链路。
    multi = extract_counterparty_trade_rows(ctx["counter_seg"], yds=ctx["yds"])
    if multi:
        return multi, bool(ctx["yds"])
    multi = extract_splits_structured(ctx["counter_seg"], yds=ctx["yds"], allow_missing_yd=allow_subject_anchor and not ctx["yds"])
    multi_from_anchor = False
    if multi:
        return multi, bool(ctx["yds"])
    multi = extract_multi(ctx["base_rec"], ctx["mkt"])
    if not multi:
        multi = extract_splits_generic(ctx["counter_seg"])
    if not multi:
        multi = extract_splits_by_yd_anchor(ctx["counter_seg"])
        multi_from_anchor = bool(multi)
    if not multi and allow_subject_anchor and not ctx["yds"]:
        multi = extract_splits_structured(ctx["counter_seg"], allow_missing_yd=True)
        multi_from_anchor = bool(multi)
    return multi, multi_from_anchor

def _append_multi_rows(ctx, multi, multi_from_anchor=False):
    rows = []
    odc, otc = other_codes(ctx["rec"], ctx["mine"], ctx["mpos"])
    odc = ctx["counter_dealer_code"] or odc
    otc = ctx["counter_trader_code"] or otc
    short = (ctx["counter_short"] or dealer_name_from_code(odc)) if ctx["mkt"] == "深交所" else ""
    shared_guoquan = ctx["guoquan"] or (dealer_name_from_code(odc) if ctx["mkt"] == "深交所" else "")
    o_person = ctx["o_person"] or person_name(ctx["rec"])
    sh_tcode = ""
    if ctx["mkt"] == "上交所":
        my_i = RE_ICODE.search(ctx["mineseg"])
        my_i = my_i.group(0) if my_i else ""
        for m in RE_ICODE.finditer(ctx["counter_seg"]):
            if m.group(0) != my_i:
                sh_tcode = m.group(0)
                break
    for item in multi:
        acct, subj, sz, yd, tcode, explicit_row_org = _multi_parts(item)
        r = _make_row(ctx, sz, yd, acct=acct, subj=subj)
        row_org = explicit_row_org or org_from_account(acct)
        r["对手方交易员"] = o_person
        r["对手方交易员代码"] = tcode or sh_tcode or otc
        if ctx["mkt"] == "深交所":
            if odc:
                r["对手方交易商代码"] = odc
            if short:
                r["对手方交易商简称"] = short
        # 交易商/通道是共享的强证据；只有公共块未给出机构时，才从该拆单账户的明确前缀推导。
        if ctx["mkt"] == "深交所" and odc:
            r["过券"] = shared_guoquan or row_org
        else:
            r["过券"] = row_org or shared_guoquan
        if not yd:
            r["备注"] = "约定号原文未给出，需人工核对补充"
        elif not sz:
            r["备注"] = "规模原文未分列，需人工确认"
        rows.append(r)
    total_yd = len(ctx["yds"])
    if total_yd and len(rows) != total_yd:
        for r in rows:
            if not r["备注"]:
                r["备注"] = "拆单笔数与约定号数量不一致，需人工核对"
    subjs = [_multi_parts(item)[1] for item in multi if _multi_parts(item)[1]]
    if subjs and len(set(subjs)) < len(subjs):
        for r in rows:
            if not r["备注"]:
                r["备注"] = "拆单内多笔命中同一交易主体代码，疑似识别错位，需人工核对"
    if not multi_from_anchor:
        anchor_check = extract_splits_by_yd_anchor(ctx["counter_seg"])
        if anchor_check:
            subj_a = {_multi_parts(item)[1] for item in multi if _multi_parts(item)[1]}
            subj_b = {_multi_parts(item)[1] for item in anchor_check if _multi_parts(item)[1]}
            if subj_a and subj_b and subj_a != subj_b:
                for r in rows:
                    if not r["备注"]:
                        r["备注"] = "与约定号锚点兜底的独立读取结果不一致，需人工核对"
    return rows


def _parse_multi_rows(ctx, allow_subject_anchor=False):
    multi, multi_from_anchor = _resolve_multi_rows(ctx, allow_subject_anchor=allow_subject_anchor)
    if not multi:
        return []
    return _append_multi_rows(ctx, multi, multi_from_anchor=multi_from_anchor)

def parse_template_single(ctx):
    rows = []
    if ctx["sizes"] and ctx["yds"] and len(ctx["sizes"]) == len(ctx["yds"]) and len(ctx["yds"]) > 1:
        for s, y in zip(ctx["sizes"], ctx["yds"]):
            rows.append(_make_row(ctx, s, y))
        return rows
    if len(ctx["yds"]) > 1:
        for y in ctx["yds"]:
            rows.append(_make_row(ctx, "", y))
        for r in rows:
            r["备注"] = "拆单-规模/对方账户需人工核对"
        return rows
    size = get_size(ctx["base_rec"], code=ctx["raw_code"], name=ctx["name"])
    rows.append(_make_row(ctx, size, ctx["yds"][0] if ctx["yds"] else ""))
    return rows

def parse_template_t3(ctx):
    if ctx["paren"]:
        rows = []
        org = pick_org(ctx["counter_seg"]) or pick_org(ctx["rec"])
        for i, (nm, sz) in enumerate(ctx["paren"]):
            acct = re.sub(r'^' + re.escape(org) + r'\s*', '', nm) if org else nm
            r = _make_row(ctx, sz, ctx["yds"][i] if i < len(ctx["yds"]) else "", acct=acct)
            r["过券"] = org or acct
            rows.append(r)
        return rows
    multi_rows = _parse_multi_rows(ctx)
    if multi_rows:
        return multi_rows
    if ctx["sizes"] and ctx["yds"] and len(ctx["sizes"]) == len(ctx["yds"]):
        return [_make_row(ctx, s, y) for s, y in zip(ctx["sizes"], ctx["yds"])]
    if len(ctx["yds"]) > 1:
        rows = [_make_row(ctx, "", y) for y in ctx["yds"]]
        for r in rows:
            r["备注"] = "拆单-规模/对方账户需人工核对"
        return rows
    return parse_template_single(ctx)

def parse_template_t4(ctx):
    multi_rows = _parse_multi_rows(ctx)
    if multi_rows:
        return multi_rows
    return parse_template_t3(ctx)

def parse_template_t5(ctx):
    multi_rows = _parse_multi_rows(ctx, allow_subject_anchor=True)
    if multi_rows:
        return multi_rows
    return parse_template_t3(ctx)

def parse_template_t6(ctx):
    rows = []
    odc, otc = other_codes(ctx["rec"], ctx["mine"], ctx["mpos"])
    odc = ctx["counter_dealer_code"] or odc
    otc = ctx["counter_trader_code"] or otc
    explicit_org = ctx["guoquan"] or dealer_name_from_code(odc)
    person = person_name(ctx["rec"])
    for t in ctx["table"]:
        r = _make_row(ctx, t['size'], t['yd'], acct=t['acct'], subj=t['subj'])
        r['行权收益率'] = t['xingquan']
        r['到期收益率'] = ''
        r['原始净价'] = t['price']
        r['交易净价'] = t['price']
        if odc:
            r['对手方交易商代码'] = odc
        if otc:
            r['对手方交易员代码'] = otc
        if person:
            r['对手方交易员'] = person
        r['过券'] = explicit_org or org_from_account(t['acct'])
        rows.append(r)
    return rows

def parse_template_t7(ctx):
    rows = []
    for b in ctx["multibond"]:
        r = _make_row(ctx, b['size'], b['yd'])
        r['债券代码'] = b['code']
        r['债券简称'] = b['name']
        r['到期收益率'] = b['yld']
        r['原始净价'] = b['price']
        r['交易净价'] = b['price']
        rows.append(r)
    return rows

def parse_template_t8(ctx):
    if _looks_like_multi_t8(ctx):
        multi_rows = _parse_multi_rows(ctx, allow_subject_anchor=True)
        if multi_rows:
            return multi_rows
    return parse_template_single(ctx)


def _line_bounds(text, start, end=None):
    end = start if end is None else end
    lo = text.rfind('\n', 0, start) + 1
    hi = text.find('\n', end)
    return lo, len(text) if hi < 0 else hi


def _size_value(num, unit):
    v = float(str(num).replace(',', ''))
    u = unit.lower().replace(' ', '')
    if u.endswith('元'):
        u = u[:-1]
    if u in ('w', '万'):
        return int(v)
    if u in ('k', 'kw'):
        return int(v * 1000)
    if u in ('e', '亿'):
        return int(v * 10000)
    return ""


def _size_line_is_noise(line, code="", name="", row_mode=False):
    if row_mode:
        return False
    if code and re.search(re.escape(code), line, re.I):
        return False
    if name and name in line:
        return False
    if re.search(r'净价|票面|到期|行权|出给|卖给|买自|\bto\b|\bfrom\b|买入|卖出|约定号', line, re.I):
        return False
    return bool(re.search(r'交易商(?:代码|号)?|交易员(?:代码|号)?|主体(?:代码|名称|简称|全称)|席位号', line))


SIZE_CTX = re.compile(r'净价|票面|到期|行权|出给|卖给|买自|\bto\b|\bfrom\b|买入|卖出|约定号|交易所', re.I)
RE_SIZE_INT = re.compile(r'(?<![A-Za-z0-9.])(\d{2,6})(?![A-Za-z0-9.])')


def _score_size_candidate(line, start, code="", name="", explicit=False, row_mode=False, near_field=False):
    score = 2 if explicit else 1
    if SIZE_CTX.search(line):
        score += 2
    anchor = line.find(code) if code else -1
    if anchor >= 0 and abs(start - anchor) <= 40:
        score += 2
    if name:
        ni = line.find(name)
        if ni >= 0 and abs(start - ni) <= 24:
            score += 1
    if near_field:
        score += 2
    if row_mode:
        score += 1
    return score


def _normalize_size_scan_text(text, code="", name=""):
    text = _norm_basic(text or "")
    for token in (code, name):
        if token:
            text = re.sub(re.escape(token), ' ', text, flags=re.I)
    # 带单位规模已在第一轮按高置信候选收集；从无单位扫描文本中移除，避免把 1,500万 再拆成 500。
    text = RE_SIZE_TOKEN.sub(' ', text)
    text = RE_YD.sub(' ', text)
    text = re.sub(r'交易商(?:代码|号)?\s*[:：]?\s*\d{6}', ' ', text)
    text = re.sub(r'交易员(?:代码|号)?\s*[:：]?\s*[0-9A-Z]{7,8}', ' ', text)
    text = re.sub(r'交易(?:商)?主体(?:代码|名称|简称|全称)?\s*[:：]?\s*36\d{8}', ' ', text)
    text = re.sub(r'36\d{8}|i\d{9}|[A-Za-z一-龥]*[Zz]\d{5,}', ' ', text)
    text = re.sub(r'20\d{2}[./-]\d{1,2}[./-]\d{1,2}|20\d{6}', ' ', text)
    text = re.sub(r'(?<!\d)\d{1,2}[./-]\d{1,2}(?!\d)', ' ', text)
    text = re.sub(r'(?:今[日天]|上交所|深交所|交易所)\s*[+＋]\d{1,2}', ' ', text)
    text = re.sub(r'\d+(?:\.\d+)?%', ' ', text)
    text = re.sub(r'净价\s*1\d{2}(?:\.\d+)?|1\d{2}(?:\.\d+)?\s*净价', ' 净价 ', text)
    text = re.sub(r'\d+\.\d+', ' ', text)
    return text


def _iter_size_candidates(rec, code="", name="", row_mode=False):
    out=[]
    for m in RE_SIZE_TOKEN.finditer(rec):
        lo, hi = _line_bounds(rec, m.start(1), m.end(2))
        line = rec[lo:hi]
        if _size_line_is_noise(line, code=code, name=name, row_mode=row_mode):
            continue
        val = _size_value(m.group(1), m.group(2))
        if val == "":
            continue
        score = _score_size_candidate(line, m.start(1) - lo, code=code, name=name, explicit=True, row_mode=row_mode)
        out.append((score, val, -m.start(1), val))

    for raw_line in rec.splitlines():
        line = raw_line.strip()
        if not line or _size_line_is_noise(line, code=code, name=name, row_mode=row_mode):
            continue
        scan_line = _normalize_size_scan_text(line, code=code, name=name)
        seen = set()
        contextual = [
            re.compile(r'(?<![.\d])(\d{2,6})(?=\s*[,，、]?\s*(?:净价|票面|到期|行权))'),
            re.compile(r'(?<![.\d])(\d{2,6})(?=\s*[,，、]?\s*约定号)'),
            re.compile(r'(?:买入|卖出)\s*[,，、]?\s*(\d{2,6})(?=\s+[,，、]?[一-龥A-Za-z0-9])'),
            re.compile(r'(?<!\d)1\d{2}(?:\.\d+)?\s+(\d{2,6})(?=\s+[,，、]?[一-龥A-Za-z])'),
        ]
        for pat in contextual:
            for m in pat.finditer(scan_line):
                val = int(m.group(1))
                if val < 50:
                    continue
                seen.add((m.start(1), val))
                score = _score_size_candidate(line, m.start(1), code=code, name=name, near_field=True, row_mode=row_mode)
                out.append((score, val, -m.start(1), val))
        for m in RE_SIZE_INT.finditer(scan_line):
            val = int(m.group(1))
            if val < 50 or (m.start(1), val) in seen:
                continue
            score = _score_size_candidate(line, m.start(1), code=code, name=name, row_mode=row_mode)
            out.append((score, val, -m.start(1), val))
    return out


def get_size(rec, code="", name=""):
    cands=_iter_size_candidates(rec, code=code, name=name)
    if cands:
        cands.sort(reverse=True)
        return cands[0][3]
    return ""

BOND_NAME_BAD_WORDS = (
    "净价", "交易", "约定", "出给", "买入", "卖出", "交易所",
    "主体", "名称", "代码", "票面", "到期", "行权", "发单", "先发", "报价", "要素已定",
)
BOND_NAME_RULES = (
    ("year_caps_cn_tail", re.compile(r'[1-9]\d[A-Z]{1,4}[一-龥]{1,8}[A-Z0-9]{1,4}'), 40),
    ("year_cn_caps_tail", re.compile(r'[1-9]\d[一-龥]{1,8}[A-Z]{1,4}\d{0,4}'), 38),
    ("year_cn_digits", re.compile(r'[1-9]\d[一-龥]{1,8}\d{2,3}'), 36),
    ("year_caps", re.compile(r'[1-9]\d[A-Z]{2,8}[A-Z0-9]{0,4}'), 34),
    ("cn_digit_cn_caps_tail", re.compile(r'[一-龥]{1,4}\d{2}[一-龥]{1,4}[A-Z]{1,3}\d{1,3}'), 33),
    ("cn_caps_digits", re.compile(r'[一-龥]{2,8}[A-Z]{1,4}\d{1,4}'), 31),
    ("cn_word_digits", re.compile(r'[一-龥]{2,8}(?:优|次|转|永续|次级)?\d{2}'), 30),
    ("generic_cn_digit", re.compile(r'[A-Za-z0-9一-龥]{4,16}'), 12),
)
RE_BOND_NAME_PIECE = re.compile(r'[A-Za-z0-9一-龥]+')
BOND_CTX = re.compile(r'净价|票面|约定号|出给|\bto\b|\bfrom\b|买入|卖出|到期|行权|交易所')


def _normalize_bond_name_candidate(text):
    nm = re.sub(r'\s+', '', _norm_basic(text or ""))
    return nm.strip('：:，,；;（）()[]【】"\'“”‘’')


def _classify_bond_name(nm):
    if not nm:
        return None
    if not 4 <= len(nm) <= 16:
        return None
    if any(word in nm for word in BOND_NAME_BAD_WORDS):
        return None
    if looks_like_account_name(nm):
        return None
    if looks_like_plain_org(nm):
        return None
    if re.fullmatch(r'00[A-Z0-9]{4,10}', nm):
        return None
    if re.fullmatch(r'20\d{6}|\d{4}', nm):
        return None
    for label, pat, score in BOND_NAME_RULES:
        if not pat.fullmatch(nm):
            continue
        if label == "generic_cn_digit" and (not re.search(r'[一-龥]', nm) or not re.search(r'\d', nm)):
            continue
        return label, score
    return None

def _find_spaced_text(rec, text, optional_hao=False):
    if not text: return None
    base=text[:-1] if optional_hao and text.endswith('号') else text
    pat=r'[ \t]*'.join(re.escape(ch) for ch in base)
    if optional_hao and text.endswith('号'):
        pat+=r'(?:[ \t]*号)?'
    return re.search(pat, rec)


def _tokenize_bond_name_pieces(text):
    return [(m.group(0), m.start(), m.end()) for m in RE_BOND_NAME_PIECE.finditer(text or "")]


def _bond_name_noise_stripped(text, code=""):
    text = _norm_basic(text or "")
    if code:
        text = re.sub(re.escape(code), ' ', text, flags=re.I)
    # Z/i 码本身是噪声，但码前中文通常是机构或账户名，不能连同“首创证券”一起吞掉。
    text = re.sub(r'i\d{9}|[A-Za-z]*[Zz]\d{5,}', ' ', text)
    text = re.sub(r'20\d{2}[./-]\d{1,2}[./-]\d{1,2}|20\d{6}', ' ', text)
    text = re.sub(r'(?<!\d)\d{1,2}[./-]\d{1,2}(?!\d)|(?<!\d)\d{4}(?=\s*[+＋]\d{1,2})', ' ', text)
    text = re.sub(r'[+＋]\d{1,2}(?!\d)', ' ', text)
    return text


def _candidate_from_pieces(pieces):
    return _normalize_bond_name_candidate(' '.join(piece for piece, _, _ in pieces))


def _iter_bond_name_windows(pieces, side=None, max_parts=3):
    if not pieces:
        return
    if side == "before":
        for size in range(1, min(max_parts, len(pieces)) + 1):
            yield pieces[-size:]
        return
    if side == "after":
        for size in range(1, min(max_parts, len(pieces)) + 1):
            yield pieces[:size]
        return
    for size in range(1, min(max_parts, len(pieces)) + 1):
        for start in range(0, len(pieces) - size + 1):
            yield pieces[start:start + size]


def _score_bond_name_candidate(nm, line, rule_score, anchor_side="", piece_count=1, fallback=False):
    score = rule_score
    if BOND_CTX.search(line):
        score += 3
    if re.search(r'交易商|交易员|主体(?:代码|名称|简称|全称)|席位号', line):
        score -= 6
    if anchor_side:
        score += 20
        if piece_count == 1:
            score += 4
        elif piece_count == 2:
            score += 2
        if anchor_side == "after":
            score += 1
    if fallback:
        score -= 8
    return score


def _collect_bond_name_candidates(rec, code=""):
    rec = _norm_basic(rec)
    seen = set()
    candidates = []

    def add_candidate(raw_nm, line, anchor_side="", piece_count=1, fallback=False):
        nm = _normalize_bond_name_candidate(raw_nm)
        classified = _classify_bond_name(nm)
        if not classified:
            return
        key = (nm, anchor_side, fallback)
        if key in seen:
            return
        seen.add(key)
        label, rule_score = classified
        score = _score_bond_name_candidate(nm, line, rule_score, anchor_side=anchor_side, piece_count=piece_count, fallback=fallback)
        candidates.append(dict(
            score=score,
            length=len(nm),
            name=nm,
            label=label,
            rule_score=rule_score,
            anchor_side=anchor_side,
            fallback=fallback,
        ))

    if code:
        m_code = re.search(re.escape(code), rec, re.I)
        if m_code:
            lo, hi = _line_bounds(rec, m_code.start(), m_code.end())
            line = rec[lo:hi]
            rel_start = m_code.start() - lo
            rel_end = m_code.end() - lo
            left = _bond_name_noise_stripped(line[:rel_start], code=code)
            right = _bond_name_noise_stripped(line[rel_end:], code=code)
            left_pieces = _tokenize_bond_name_pieces(left)
            right_pieces = _tokenize_bond_name_pieces(right)
            for win in _iter_bond_name_windows(right_pieces, side="after"):
                add_candidate(_candidate_from_pieces(win), line, anchor_side="after", piece_count=len(win))
            for win in _iter_bond_name_windows(left_pieces, side="before"):
                add_candidate(_candidate_from_pieces(win), line, anchor_side="before", piece_count=len(win))
            if candidates:
                candidates.sort(key=lambda x: (-x["score"], -x["length"], x["name"]))
                return candidates

    for raw_line in rec.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        cleaned = _bond_name_noise_stripped(line, code=code)
        pieces = _tokenize_bond_name_pieces(cleaned)
        for win in _iter_bond_name_windows(pieces, max_parts=3):
            add_candidate(_candidate_from_pieces(win), line, fallback=True, piece_count=len(win))
    candidates.sort(key=lambda x: (-x["score"], -x["length"], x["name"]))
    return candidates


def _best_bond_name_candidate(rec, code=""):
    candidates = _collect_bond_name_candidates(rec, code=code)
    return candidates[0] if candidates else None


def has_confident_bond_name(rec, code=""):
    cand = _best_bond_name_candidate(rec, code=code)
    if not cand:
        return False
    return cand["rule_score"] >= 30


def get_name(rec, code=""):
    cand = _best_bond_name_candidate(rec, code=code)
    return cand["name"] if cand else ""

def dealer_org(seg):
    m=RE_DEALER_ORG.search(seg)
    return m.group(1) if m else ""

# 旧专用多拆规则：这些正则保留为 legacy fallback，只在统一结构块解析没命中时兜底。
# 目标不是继续往这里堆样本，而是逐步把稳定场景前移到 extract_splits_structured。
# 返回 [(对方账户,主体代码,规模万,约定号,交易员码)] ；<2 个则空
# 上交所多拆：逐行"产品名(可无号) [可选i码] 规模(单位可选) 约定号"；规模须空格分隔+紧跟约定号，避免误匹配i码/单笔
RE_MULTI_SH = re.compile(r'(?:^|\n)[ \t]*([一-龥A-Za-z0-9][^\n]*?)(?:i\d{9})?[ \t（(]+(\d{1,6})\s*([wWkK万]?)[）)]?\s*[；;，, ]*约定号[：:\s]*(\d+)', re.M)
# 上交所多拆(规模在前)：如 "1）1500万 广东省拾贰号职业年金计划海富通组合 约定号842"
RE_MULTI_SH2= re.compile(r'(?:^|\n)\s*(?:\d{1,2}\s*[）)、.]\s*)?(\d{2,6})\s*([wWkK万])\s*([一-龥A-Za-z0-9（）()－\-]+?)\s*约定号[：:\s]*(\d+)', re.M)
RE_MULTI_SH3= re.compile(r'^(.*?)\s+(\d{2,6})\s*([wWkK万]?)(?:\s+([A-Za-z一-龥]*[Zz]\d{5,}|i\d{9}))?\s*$', re.M)
RE_MULTI_SZ = re.compile(r'(36\d{8})\s*([^\d\s：:][^\n]*?)\s*(?:[（(]?\s*(\d+)\s*[wW万][）)]?)?\s*约定号[：:\s]*(\d+)')  # 主体码 名 [规模万] 约定号
RE_SUBJ_LABEL = re.compile(r'^(?:交易)?(?:商)?主体(?:名称|简称|代码)?\s*[：:]\s*')  # 去名字前的标签
RE_MULTI_SZ2= re.compile(r'([一-龥A-Za-z0-9]{2,16}号)\s*(\d+)\s*[wW万]\s*约定号\s*(\d+)')
RE_MULTI_FZ = re.compile(r'([一-龥]{2,4}\d{1,2})\s*(\d+)\s*万[；;，, ]*约定号\s*(\d+)')   # 玺福19 500万 约定号18420842
RE_SUBJ_NAME= re.compile(r'(36\d{8})\s*[^0-9]{0,4}?([一-龥A-Za-z0-9\-（）]+号)')            # 主体代码+全名
RE_MULTI_B  = re.compile(r'([一-龥][一-龥A-Za-z0-9]*)\s+(\d+)\s*([wWkK万])\s+(36\d{8})\s+约定号[：:\s]*(\d+)')  # 名 规模单位 主体码 约定号(中邮式)
RE_MULTI_A  = re.compile(r'(36\d{8})\s+([一-龥][一-龥A-Za-z0-9]*号)\s+(\d+)\s+[一-龥][^\n约]*?约定号[：:\s]*(\d+)')  # 主体码 名 规模 名 约定号(安阳式)
def extract_multi(seg, mkt):
    out=[]
    if mkt=='上交所':
        for m in RE_MULTI_SH.finditer(seg):        # 产品名在前
            v=int(m.group(2)); u=m.group(3).lower()
            sz=v*1000 if u=='k' else v
            nm=RE_SUBJ_LABEL.sub('', m.group(1)).strip(' \t：:，,；;')
            out.append((nm,'',sz,m.group(4),''))
        if len(out)<2:                              # 规模在前(如 1）1500万 广东省…组合 约定号842)
            for m in RE_MULTI_SH2.finditer(seg):
                v=int(m.group(1)); u=m.group(2).lower()
                sz=v*1000 if u=='k' else v
                nm=RE_SUBJ_LABEL.sub('', m.group(3)).strip(' \t：:，,；;（(').strip()
                out.append((nm,'',sz,m.group(4),''))
    else:
        for m in RE_MULTI_B.finditer(seg):      # 名 规模单位 主体码 约定号(中邮式)
            v=int(m.group(2)); u=m.group(3).lower()
            out.append((m.group(1), m.group(4), v*1000 if u=='k' else v, m.group(5), ''))
        if len(out)<2:                          # 主体码 名 规模 名 约定号(安阳式)
            out=[]
            for m in RE_MULTI_A.finditer(seg):
                out.append((m.group(2), m.group(1), int(m.group(3)), m.group(4), ''))
        if len(out)<2:                          # 通用：主体码 名 [规模万] 约定号
            out=[]
            for m in RE_MULTI_SZ.finditer(seg):
                sz=int(m.group(3)) if m.group(3) else ''
                nm=extract_full_account_from_line(RE_SUBJ_LABEL.sub('', m.group(2)))
                nm,sz=split_trailing_size(nm,sz)      # 名尾若跟"规模区间整数"则切出来当规模
                out.append((nm,m.group(1),sz,m.group(4),''))
        if len(out)<2:   # 方正式：短名+规模万+约定号 分列，主体代码/全称另行分列(标签不一)
            fz=RE_MULTI_FZ.findall(seg)
            if len(fz)>=2:
                # 全称按位置配"其前最近的36主体代码"
                codes=[(m.start(),m.group(1)) for m in re.finditer(r'(36\d{8})', seg)]
                names=[(m.start(),m.group(1)) for m in re.finditer(r'(?:全称|名称)[：:]?\s*([一-龥][^；;\n，]*?号)', seg)]
                full2code={}
                for npos,nm in names:
                    prev=[c for cpos,c in codes if cpos<npos]
                    if prev: full2code[nm]=prev[-1]
                out=[]
                for short,size,yd in fz:
                    subj=''; acct=short+'号'
                    for full,c in full2code.items():
                        if short in full:
                            subj=c; acct=full; break
                    out.append((acct,subj,int(size),yd,''))
        if len(out)<2:   # 主体码+名(规模可缺)，约定号在末尾"+"连写(华安/东吴式)；名须含"号"以排除我方主体
            prods=[(c,n) for c,n in re.findall(r'(36\d{8})\s*([一-龥][^\n]*)', seg)
                   if '号' in n and '交易主体名称' not in n]
            szs =re.findall(r'(36\d{8})\s*[一-龥][^\n]*?(\d+)\s*[wW万]', seg)  # 若带规模单位则取之
            szmap={c:int(s) for c,s in szs}
            ydm=re.search(r'约定号[：:\s]*([0-9]+(?:[+＋][0-9]+)+)', seg)
            if len(prods)>=2 and ydm:
                yds=re.split(r'[+＋]', ydm.group(1))
                if len(yds)==len(prods):
                    out=[(RE_SUBJ_LABEL.sub('',n).strip(' \t：:，,；;（('), c, szmap.get(c,''), yds[i], '')
                         for i,(c,n) in enumerate(prods)]
    # 规模在独立行(如临动GT01: 2000\n交易商代码…\n约定号)——产品都缺规模时，按顺序配独立数字行
    if out and all(not x[2] for x in out):
        szl=re.findall(r'(?:^|\n)[ \t]*(\d{3,6})[ \t]*万?[ \t]*(?=\n)', seg)
        if len(szl)>=len(out):
            out=[(a,b,int(szl[i]),d,e) for i,(a,b,c,d,e) in enumerate(out)]
    return out if len(out)>=2 else []

# 表格式(5+5 行权)：日期 期限 简称 对方 代码 规模 行权 收益率 净价 我方 主体代码 约定号
RE_TABLE = re.compile(r'20\d{6}\s+\S+\s+\S+\s+([一-龥A-Za-z0-9－\-]+)\s+\d{6}\.[A-Z]{2}\s+(\d+)\s+行权\s+([\d.]+)\s+(\d+)\s+\S+\s+(36\d{8})\s+(70\d{5,7})')
def extract_table(rec):
    out=[]
    for m in RE_TABLE.finditer(rec):
        out.append(dict(acct=m.group(1),size=int(m.group(2)),xingquan=m.group(3),
                        price=m.group(4),subj=m.group(5),yd=m.group(6)))
    return out if len(out)>=2 else []

def split_trailing_size(name, cur_size):
    """名字尾部若跟着一个'规模区间整数'(带空格、不粘号)，按数值切出来当规模。
    数值区间约定：小数/≈100=净价或收益率(不当规模)；整数且≥50=规模(万)。"""
    if cur_size!='' or not name: return name, cur_size
    m=re.match(r'^(.*[一-龥号])\s+(\d{2,6})\s*([wWkK万]?)$', name.strip())
    if m:
        v=int(m.group(2)); u=(m.group(3) or '').lower()
        if u or v>=50:            # 有单位，或数值达到规模区间
            return m.group(1).strip(), (v*1000 if u=='k' else v)
    return name, cur_size

def _gen_prodname(line):
    """多笔行优先按字段边界读取完整账户，旧词形规则只作兜底。"""
    l=line
    for w in WHITELIST: l=l.replace(w,' ')
    l=RE_MINE_PAT.sub(' ', l)
    l=re.sub(r'(交易主体名称|交易主体简称|主体名称|交易商名称)[：:]\s*','',l)
    full = extract_full_account_from_line(l)
    if full and looks_like_account_name(full) and not mine_related_text(full):
        return full
    compound_tail = r'(?:\d{1,3}个月)?(?:集合资产管理计划|资产管理计划|职业年金计划|企业年金计划|年金计划)?'
    for m in re.finditer(r'[一-龥][一-龥A-Za-z0-9－\-（）]*?(?:\d+M?\d*号' + compound_tail + r'|\d+号' + compound_tail + r'|集合资产管理计划|职业年金计划|企业年金计划|年金|组合)', l):
        nm=RE_SUBJ_LABEL.sub('', m.group(0)).strip('－- ')
        if nm and not RE_NOTNAME.match(nm): return nm
    return ''
def _gen_size(line):
    cands = _iter_size_candidates(line, row_mode=True)
    if not cands:
        return ''
    cands.sort(reverse=True)
    return cands[0][3]
RE_LEAD_ENUM = re.compile(r'^[ \t]*(?:\d{1,2}\s*[.、）)．]|[①②③④⑤⑥⑦⑧⑨⑩]|[a-zA-Z]\s*[.、)])\s*')
def _strip_leading_enum(line):
    """去掉行首编号/项目符号（1. / a) / ① 等），这类前缀在 T4/T5/T8 的清单块里反复出现，
    属于跟哪个模板无关的通用噪声，不应该每条正则各自兼容一遍。"""
    return RE_LEAD_ENUM.sub('', line)


def _line_trade_code(line):
    m = RE_ICODE.search(line) or re.search(r'([Zz]\d{5,})', line)
    if not m:
        return ""
    return m.group(1).upper() if m.lastindex else m.group(0)


def _multi_parts(item):
    """兼容旧的5字段拆单结果和新的6字段逐行结果。"""
    if len(item) >= 6:
        return item[:6]
    acct, subj, size, yd, tcode = item
    return acct, subj, size, yd, tcode, ""


def _execution_org_from_repeated_bond_line(line, tcode=""):
    """取重复债券明细行中的执行机构；短称必须与Z/i码相邻才入选。"""
    text = _norm_basic(line or "")
    price = re.search(r'(?:净价|成交净价|交易净价)\s*[:：]?\s*\d+(?:\.\d+)?', text)
    tail = text[price.end():] if price else text
    orgs = [normalize_visible_org(m.group(1)) for m in RE_ORG.finditer(tail)]
    orgs = [org for org in orgs if org and not mine_related_text(org)]
    if orgs:
        return orgs[-1]
    if not tcode:
        return ""
    code_match = re.search(re.escape(tcode), tail, re.I)
    before_code = tail[:code_match.start()] if code_match else tail
    short_match = re.search(r'([一-龥]{2,20})\s*$', before_code)
    if not short_match:
        return ""
    candidate = short_match.group(1)
    if re.search(r'号|计划|组合|基金|年金|产品|私募', candidate):
        return ""
    return candidate


def extract_counterparty_trade_rows(seg, yds=None):
    """
    逐行交易锚点：对方段内的“账户 + 规模 + 机构”各自成笔。

    债券代码可以在每行重复，也可以只写在公共头中；方括号、冒号等只做清洗，
    不参与模板判定。返回值比旧拆单结果多一个row_org，用于保留每行自己的过券机构。
    """
    parsed = []
    for raw in (seg or "").splitlines():
        line = _strip_leading_enum(raw).strip()
        bond = RE_BOND.search(line)
        size_anchor = RE_SIZE_TOKEN.search(line)
        if not size_anchor:
            continue
        account_end = bond.start() if bond and bond.start() < size_anchor.start() else size_anchor.start()
        account = _clean_counterparty_candidate(line[:account_end]).strip(' 　\t：:，,；;（）()')
        size = _gen_size(line)
        tcode = _line_trade_code(line)
        row_org = _execution_org_from_repeated_bond_line(line, tcode=tcode)
        if not account or not size or not row_org:
            continue
        parsed.append((account, '', size, '', tcode, row_org))
    if len(parsed) < 2:
        return []
    yds = list(yds or [])
    return [
        (acct, subj, size, yds[i] if i < len(yds) else '', tcode, row_org)
        for i, (acct, subj, size, _, tcode, row_org) in enumerate(parsed)
    ]


def _slice_structured_counterparty_blocks(seg, yds=None):
    text = _norm_basic(seg or "")
    positions = _structured_counterparty_anchor_positions(text)
    if len(positions) < 2:
        return []
    tail_end = len(text)
    if yds:
        shared_yd = re.search(r'约(?:定号)?[：:\s]*\d+(?:\s*[+＋、，,]\s*\d+)+', text)
        if shared_yd:
            tail_end = shared_yd.start()
    blocks = []
    for i, start in enumerate(positions):
        end = positions[i + 1] if i + 1 < len(positions) else tail_end
        block = text[start:end].strip()
        if block:
            blocks.append(block)
    return blocks if len(blocks) >= 2 else []


def _parse_structured_counterparty_block(block):
    block = _norm_basic(block or "").strip()
    if not block:
        return None
    name = extract_subject_name(block) or _gen_prodname(block)
    subj = ""
    m_subj = re.search(r'(36\d{8})', block)
    if m_subj:
        subj = m_subj.group(1)
    trader = ""
    m_trader = RE_TRADER.search(block) or RE_TRADER2.search(block)
    if m_trader:
        trader = m_trader.group(1)
    size = _gen_size(block)
    tcode = _line_trade_code(block)
    if not (name or subj):
        return None
    return dict(acct=name, subj=subj, size=size, trader=trader, tcode=tcode)


def extract_splits_structured(seg, yds=None, allow_missing_yd=False):
    """统一的主体块解析器。
    优先按重复出现的结构锚点切块（规模独立行 / 交易主体标签行 / 主体码行），
    再块内统一取 账户、主体码、规模、交易员码。这样模板识别依赖结构，不依赖样本文案。"""
    blocks = _slice_structured_counterparty_blocks(seg, yds=yds)
    if len(blocks) < 2:
        return []
    parsed = []
    for block in blocks:
        item = _parse_structured_counterparty_block(block)
        if not item:
            return []
        parsed.append(item)
    if yds:
        if len(parsed) != len(yds):
            return []
        return [(item["acct"], item["subj"], item["size"], yds[i], item["trader"] or item["tcode"]) for i, item in enumerate(parsed)]
    if not allow_missing_yd:
        return []
    return [(item["acct"], item["subj"], item["size"], "", item["trader"] or item["tcode"]) for item in parsed]


def _extract_sh_account_rows(seg):
    rows = []
    for raw in (seg or "").splitlines():
        line = _strip_leading_enum(raw).strip()
        if not line or RE_YD_LINE.search(line):
            continue
        if RE_BOND.search(line) or re.search(r'净价|行权|到期', line):
            continue
        if re.search(r'交易商(?:代码|号)?|交易员(?:代码|号)?|交易主体(?:代码|名称|简称|全称)|席位号', line):
            continue
        m = RE_MULTI_SH3.match(line)
        if not m:
            continue
        name = re.sub(r'^(?:出给|to|from|买入方?|卖出方?|买方|卖方)\s*', '', m.group(1), flags=re.I)
        name = name.strip(' \t：:，,；;（）()')
        if not name:
            continue
        if looks_like_plain_org(name) and not re.search(r'\d|号|计划|组合|基金|年金|资管|产品|私募', name):
            continue
        size = int(m.group(2))
        unit = (m.group(3) or "").lower()
        if unit == 'k':
            size *= 1000
        rows.append((name, '', size, '', _line_trade_code(line)))
    return rows


def extract_splits_generic(seg):
    """通用兜底拆单：以约定号为锚，一个约定号一笔；每笔就近取 主体码/对方名/规模(仅显式单位)。"""
    lines=[_strip_leading_enum(l) for l in seg.split('\n') if l.strip()]
    # A) 集中式：约定号 A+B+C…（产品分散在各行），按序配对
    mj=re.search(r'约(?:定号)?[：:\s]*(\d+(?:\s*[+＋、，,]\s*\d+)+)', seg)
    if mj:
        yds=[y.strip() for y in re.split(r'[+＋、，,]', mj.group(1))]
        prods=[]
        for l in lines:
            if re.search(r'约(?:定号)?[：:\s]*\d+\s*[+＋]', l): continue   # 跳过集中约定号那行
            subj=re.search(r'(36\d{8})', l); nm=_gen_prodname(l)
            if subj or nm:
                prods.append((nm, subj.group(1) if subj else '', _gen_size(l), '', _line_trade_code(l)))
        if len(prods) < 2:
            prods = _extract_sh_account_rows(seg)
        if len(prods) >= 2:
            return [
                (acct, subj, size, yds[i] if i < len(yds) else '', tcode)
                for i, (acct, subj, size, _, tcode) in enumerate(prods)
            ]
    # B) 逐行式：每行各带自己的约定号
    per=[]
    for l in lines:
        ym=re.search(r'约(?:定号)?[：:\s]*(\d+)(?![+＋\d])', l)
        if not ym: continue
        subj=re.search(r'(36\d{8})', l)
        per.append((_gen_prodname(l), subj.group(1) if subj else '', _gen_size(l), ym.group(1), ''))
    return per if len(per)>=2 else []


def _scan_block_fields(block):
    """块内不认字段顺序、不认标签具体措辞，只按字段"长相"扫描：主体码/对方名/规模/交易商代码/交易员代码。
    这是所有"多笔"模板(T3/T4/T5/T8)共享的底层不变量——一个约定号对应一段字段，
    字段本身可以用统一的"长相"规则找到，不需要为每种标签排列各写一条专用正则。"""
    subj=''
    m=re.search(r'(36\d{8})', block)
    if m: subj=m.group(1)
    dealer=''
    m=RE_DEALER.search(block)
    if m: dealer=m.group(1)
    trader=''
    m=RE_TRADER.search(block) or RE_TRADER2.search(block)
    if m: trader=m.group(1)
    return _gen_prodname(block), subj, _gen_size(block), dealer, trader

def extract_splits_by_yd_anchor(seg):
    """通用兜底拆单(第二层)：以约定号为唯一锚点切块，块内不分先后顺序扫描字段。
    仅在专用正则(RE_MULTI_*)和逐行式兜底(extract_splits_generic)都没命中时才启用，
    用来接住样本外没见过的标签排列/命名方式，同时不影响已经调好的样本内匹配。"""
    stripped='\n'.join(_strip_leading_enum(l) for l in seg.split('\n'))
    matches=[m for m in RE_YD.finditer(stripped) if not re.search(r'[+＋、，,]', m.group(1))]
    if len(matches)<2: return []
    out=[]; prev_end=0
    for m in matches:
        block=stripped[prev_end:m.end()]
        prev_end=m.end()
        name,subj,size,dealer,trader=_scan_block_fields(block)
        if not (subj or name): return []   # 块内连主体信息都没有，说明不是这种结构，整体放弃
        out.append((name,subj,size,m.group(1),trader))
    return out

def extract_all_yds(rec):
    """逐行提取真实约定号，避免把'交易主体代码/简称/量/约定号：'这类表头误当成约定号值。"""
    yds=[]
    for line in rec.splitlines():
        line=_strip_leading_enum(line)
        for m in RE_YD_LINE.finditer(line):
            yds.extend(x.strip() for x in re.split(r'[+＋、，,]', m.group(1)) if x.strip())
    return yds

def extract_paren_products(rec):
    """括号内'+'连写的产品串：(西藏信托 长盈稳健25号1200+西藏信托 善盈长盈稳健42号1600+...) → [(名,规模)]"""
    m=re.search(r'[（(]([^）)]*[一-龥][^）)]*[+＋][^）)]*)[）)]', rec)
    if not m: return []
    out=[]
    for it in re.split(r'[+＋]', m.group(1)):
        mm=re.match(r'\s*([一-龥].*?)\s*(\d{2,5})\s*$', it)
        if not mm: return []          # 有一项不符合(名+规模)就整体放弃
        out.append((mm.group(1).strip(), int(mm.group(2))))
    return out if len(out)>=2 else []

def extract_multibond(rec):
    """同一对手、多只券（每只券各带规模/净价/收益率/约定号）→ 逐券一行"""
    codes=[(m.start(),m.group(0),_canonical_bond(m)) for m in RE_BOND.finditer(rec)]
    if len(codes)<2: return []
    out=[]
    for idx,(pos,raw_code,code) in enumerate(codes):
        end=codes[idx+1][0] if idx+1<len(codes) else len(rec)
        chunk=rec[pos:end]
        ydm=RE_YD.search(chunk)
        if not ydm: return []          # 不是"每券各自约定号"结构
        ym=RE_YIELD.search(chunk)
        pm=RE_PRICE.search(chunk) or RE_PRICE2.search(chunk.replace(raw_code,''))
        name=get_name(chunk,raw_code)
        out.append(dict(code=code,name=name,size=get_size(chunk, code=raw_code, name=name),
            yld=(ym.group(1)+'%') if ym else '',price=pm.group(1) if pm else '',yd=ydm.group(1)))
    return out if len(out)>=2 else []


def _today_iso():
    return date.today().isoformat()


# 日期候选按优先级尝试：显式带年份的优先，其余(交易所后缀/月日/紧凑MMDD/裸M-D)按年内日期兜底，
# 年份用当年，不再写死；每个候选都做月1-12、日1-31的合法性校验，避免把约定号误当成紧凑日期。
_DATE_CANDIDATES = (
    (RE_DATEF, True),   # (年,月,日)
    (RE_DATEZH, True),  # 中文年月日，如 "2026年7月8日"
    (RE_DATEC, True),   # 紧凑 YYYYMMDD，如表格式 "20260703"
    (RE_DATES, False),  # "7.2交易所"/"7-2交易所"/"7/2交易所"
    (RE_DATEM, False),  # "7月2日"
    (RE_DATE4T, False), # 紧凑 "0702交易"
    (RE_DATE4P, False), # 紧凑 "0511+0"
    (RE_DATES2, False), # 裸 "6/30"/"4/14"（无交易所后缀，如 "6/30+0"）
)


def norm_date(rec):
    year = date.today().year
    for pat, has_year in _DATE_CANDIDATES:
        m = pat.search(rec)
        if not m:
            continue
        if has_year:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        else:
            y, mo, d = year, int(m.group(1)), int(m.group(2))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            continue
    if re.search(r'(?:今日|今天)(?:\s*交易所)?', rec):
        return _today_iso()
    return ""


def split_main_and_supplement(rec):
    m = RE_T8_MARKER.search(rec or "")
    if not m:
        return (rec or "").strip(), ""
    return rec[:m.start()].strip(), rec[m.start():].strip()


def mine_related_text(text):
    compact=_compact_text(text)
    if not compact:
        return False
    if any(alias in compact for alias in MINE_ALIASES):
        return True
    # 裸的我方机构前缀(如"中信信托"，不带具体产品名)本身就是我方主体，不该被当成对方机构名——
    # MINE_ALIASES 只收了"前缀+具体产品名"的组合，裸前缀单独出现时要在这里兜底
    return any(compact == prefix for prefix in MINE_PREFIXES)

def find_mine(rec):
    """返回 (我方账户, 命中位置) ；没命中返回(None,-1)"""
    labeled, labeled_pos = extract_role_account(rec, "mine", with_pos=True)
    if labeled:
        return labeled, labeled_pos
    best=None; pos=-1
    for alias in MINE_ALIASES:
        m=_find_spaced_text(rec, alias, optional_hao=alias.endswith('号'))
        if m and (pos<0 or m.start()<pos):
            best,pos=MINE_ALIAS_TO_CANON[alias],m.start()
    if best: return best,pos
    # 兜底2：模式识别我方产品（中信/粤财/财信信托…号），对方多为银行/基金/券商/年金，不会误匹配
    m=RE_MINE_PAT.search(rec)
    if m: return _compact_text(m.group(1)), m.start()
    return best,pos

def direction_and_split(rec, mine_pos):
    """判断方向，返回(方向, 我方段, 对方段)"""
    labeled_direction = RE_DIRECTION_LABEL.search(rec)
    explicit_direction = labeled_direction.group(1) if labeled_direction else ""
    to_match = re.search(r'出给|卖给|\bto\b', rec, re.I)
    from_match = re.search(r'买自|\bfrom\b', rec, re.I)

    if not explicit_direction and to_match:
        explicit_direction = "卖出" if mine_pos < 0 or mine_pos < to_match.start() else "买入"
    if not explicit_direction and from_match:
        explicit_direction = "买入" if 0 <= mine_pos < from_match.start() else "卖出"
    if not explicit_direction:
        j=rec.find("买入")
        if j>=0 and rec[j:j+3]!="买入方":
            explicit_direction = "卖出" if mine_pos > j else "买入"

    # 显式买卖方标签：卖方/卖出方、买方/买入方
    ms=re.search(r'卖(?:出)?方', rec); mb=re.search(r'买(?:入)?方', rec)
    if ms and mb:
        lo,hi=sorted([ms.start(), mb.start()])
        seg1,seg2=rec[lo:hi],rec[hi:]
        # 明确买卖方块出现时，我方账户实际落在哪个块是最强证据，优先级高于抬头中的 to/from。
        if lo <= mine_pos < hi:
            direction = "卖出" if "卖" in seg1[:4] else "买入"
            return direction, seg1, seg2
        if mine_pos >= hi:
            direction = "卖出" if "卖" in seg2[:4] else "买入"
            return direction, seg2, seg1
        if explicit_direction == "卖出":
            mineseg,otherseg = (seg1,seg2) if "卖" in seg1[:4] else (seg2,seg1)
            return "卖出", mineseg, otherseg
        if explicit_direction == "买入":
            mineseg,otherseg = (seg1,seg2) if "买" in seg1[:4] else (seg2,seg1)
            return "买入", mineseg, otherseg
        if lo<=mine_pos<hi: mineseg,otherseg=seg1,seg2
        else: mineseg,otherseg=seg2,seg1
        direction="卖出" if "卖" in mineseg[:4] else "买入"
        return direction, mineseg, otherseg
    # 分隔符 出给/to/卖给：左侧是卖方，右侧是买方。
    if to_match:
        left, right = rec[:to_match.start()], rec[to_match.end():]
        mine_left = mine_pos < 0 or mine_pos < to_match.start()
        if labeled_direction:
            mine_left = explicit_direction == "卖出"
        return (explicit_direction or ("卖出" if mine_left else "买入"), left, right) if mine_left else (explicit_direction or "买入", right, left)
    # from/买自：左侧是买方，右侧是卖方；根据我方账户位置决定方向，不再固定判为卖出。
    if from_match:
        left, right = rec[:from_match.start()], rec[from_match.end():]
        mine_left = 0 <= mine_pos < from_match.start()
        if mine_pos < 0 and labeled_direction:
            mine_left = explicit_direction == "买入"
        return (explicit_direction or ("买入" if mine_left else "卖出"), left, right) if mine_left else (explicit_direction or "卖出", right, left)
    if labeled_direction:
        return explicit_direction, rec, rec
    # 对方买入 …… 我方(在后)：如"金融街证券 i… 买入 bond … 中信信昱11号" → 我方卖出，对方在买入前
    j=rec.find("买入")
    if j>=0 and rec[j:j+3]!="买入方":
        return ("卖出", rec[j+2:], rec[:j]) if mine_pos>j else ("买入", rec[:j], rec[j+2:])
    # 只有卖方/卖家先发标记(无买方、无出给)：我方=卖方，对方靠拆单/机构映射补
    if re.search(r'卖方|卖家先发|卖出先发|卖出', rec):
        return "卖出", rec, rec
    # 兜底：整段完全没有买卖方向词(深交所有些记录只标"XX证券 发"/"我方发"这种发起方，没有买入/卖出字样)。
    # 这类记录里通常有一个固定的"我方经纪主体"说明块(交易商代码+交易员代码+交易主体代码+机构经纪名称)，
    # 但这个块不一定在文本末尾——对方信息可能写在它前面，也可能写在它后面(甚至前后都有)。
    # 所以不能简单"块之前当对方、块之后当我方"二分；而是把这个块本身挖掉当我方段，
    # 剩下的前后两截拼起来当对方段，不管对方信息落在块前还是块后都不会丢。
    m_self = RE_SELF_BROKER_BLOCK.search(rec) or re.search(r'交易(?:商)?主体(?:全称|名称)\s*[：:]\s*[一-龥]{2,16}机构经纪', rec)
    if m_self:
        mineseg = rec[m_self.start():m_self.end()]
        otherseg = rec[:m_self.start()] + rec[m_self.end():]
        return "", mineseg, otherseg
    return "", rec, ""

def other_codes(rec, mine, mine_pos):
    """取"非我方"的交易商/交易员代码：我方代码紧跟白名单之后，其余即对手方"""
    tail = rec[mine_pos: mine_pos+120] if (mine_pos is not None and mine_pos>=0) else ""
    md = RE_DEALER.search(tail); md = md.group(1) if md else ""
    mt = RE_TRADER.search(tail); mt = mt.group(1) if mt else ""
    ds=[m.group(1) for m in RE_DEALER.finditer(rec) if m.group(1)!=md]
    ts=[m.group(1) for m in RE_TRADER.finditer(rec) if m.group(1)!=mt]
    if not ts:   # 兜底1：名在码前（交易员：赵越 00H00010）
        ts=[m.group(1) for m in RE_TRADER2.finditer(rec) if m.group(1)!=mt]
    if not ts:   # 兜底2：括号内的交易员代码（毕晓韵 （000S0008））
        ts=[c for c in re.findall(r'[（(]([0-9A-Z]{7,8})[)）]', rec) if c!=mt]
    return (ds[0] if ds else ""), (ts[0] if ts else "")

RE_NOTNAME=re.compile(r'券|经纪|交易|证券|机构|资管|基金|信托|银行|资产|公司|约定|集合|计划|组合|年金|产品|管理|保险|养老|席位|主体|代码|名称|净价|类型|资本|银|买|卖|发|出给|存续|持有')
def person_name(seg):
    """对手方交易员姓名：名可在码前/码后/单独'交易员：名'，排除'…机构经纪交易员'等非人名"""
    pats=[
        r'交易员\s*(?:代码|号)?\s*[：:]?\s*i\d{9}\s+([一-龥]{2,4})(?![一-龥])',            # 上交所i码后名：i029904205 孙磊
        r'交易员[号代码]*\s*[：:]?\s*[0-9A-Z]{7,8}\s*[，,、\s]+([一-龥]{2,4})(?![一-龥])',  # 码 后名：04SE0003 周玲玲 / 00IX0005 郝爽(无冒号)
        r'交易员(?:及交易员代码|代码|名称|号)?\s*[：:]?\s*([一-龥]{2,4})\s*[（(]?\s*[0-9A-Z]{7,8}',  # 名 后码：毕晓韵（000S0008）
        r'交易员\s*名称\s*[：:]\s*([一-龥]{2,4})(?![一-龥])',                                  # 交易员名称：张锐
        r'交易员\s*[：:]\s*([一-龥]{2,4})(?![一-龥0-9A-Z])',                                   # 交易员：谭骅
        r'i\d{9}\s+([一-龥]{2,4})(?![一-龥])',                                                # 无标签 i码后名：i020038603 吴静静
        r'交易员代码\s*[:：]?\s*[0-9A-Z]{7,8}([一-龥]{2,4})(?![一-龥])',                        # 码紧贴名：006V0014蒋佳玮
    ]
    for p in pats:
        for m in re.finditer(p, seg, re.I):
            nm=m.group(1)
            if not RE_NOTNAME.search(nm): return nm
    m=re.search(r'[（(]([一-龥]{2,4})[)）]', seg)
    if m and not RE_NOTNAME.search(m.group(1)): return m.group(1)
    return ""

def pick_prod(seg):
    return counterparty.pick_prod(seg, mine_related_text)


def resolve_counterparty_account(labeled="", subject_name="", product="", head=""):
    return counterparty.resolve_counterparty_account(
        labeled, subject_name, product, head, mine_related=mine_related_text
    )


def org_from_account(account):
    return counterparty.org_from_account(account, mine_related_text)


def resolve_guoquan(seg, dealer_code="", subject_name="", head="", account="", execution_org=""):
    return counterparty.resolve_guoquan(
        seg, dealer_code, subject_name, head, account, execution_org, mine_related=mine_related_text
    )


def extract_counterparty_head(seg):
    return counterparty.extract_counterparty_head(seg, mine_related_text)


def infer_market(rec, code=""):
    if code.endswith(".SH"):
        return "上交所"
    if code.endswith(".SZ"):
        return "深交所"
    if "深交所" in rec:
        return "深交所"
    if "上交所" in rec:
        return "上交所"
    return ""


def write_rows_to_xlsx(rows, out_path):
    import openpyxl
    from openpyxl.styles import Font,PatternFill

    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="脚本提取"
    ws.append(OUTPUT_COLUMNS)
    for c in ws[1]:
        c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="305496")
    for r in rows:
        ws.append([r.get(c,"") for c in OUTPUT_COLUMNS])
    wb.save(out_path)

def main():
    base=os.path.dirname(os.path.abspath(__file__))
    files=sorted(glob.glob(os.path.join(base,"代投代缴-*.txt")))
    allrows=[]
    for f in files:
        txt=open(f,encoding='utf-8').read()
        allrows.extend(parse_text(txt))
    out=os.path.join(base,"现券预录单要素_脚本输出.xlsx")
    write_rows_to_xlsx(allrows, out)
    print("行数:",len(allrows),"->",out)
    # 命中白名单统计
    miss=sum(1 for r in allrows if r["我方账户"].startswith("(未"))
    print("未命中白名单行:",miss)

if __name__=="__main__":
    main()
